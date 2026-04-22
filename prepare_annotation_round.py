#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Zero-config annotation prepare script.

What it does:
1. Reads fresh smart labels CSV:
   annotations/labels_template_smart.csv
2. If old filled annotations exist, merges old labels into the new template by:
   - vacancy_id
   - resume_id
3. Writes:
   - annotations/labels_template_smart_merged.csv
   - annotations/labels_template_smart_merged.xlsx
4. Adds Excel drop-downs for label columns (0,1,2)

Run:
    py prepare_annotation_round.py
"""

from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_DIR = Path(".").resolve()
ANNOTATIONS_DIR = PROJECT_DIR / "annotations"

NEW_CSV = ANNOTATIONS_DIR / "labels_template_smart.csv"
OLD_FILLED_CSV = ANNOTATIONS_DIR / "labels_template_smart_filled.csv"
MERGED_CSV = ANNOTATIONS_DIR / "labels_template_smart_merged.csv"
MERGED_XLSX = ANNOTATIONS_DIR / "labels_template_smart_merged.xlsx"

KEY_COLUMNS = ["vacancy_id", "resume_id"]
ANNOTATION_COLUMNS = ["annotator_1_label", "annotator_2_label", "final_label", "comment"]
LABEL_COLUMNS = ["annotator_1_label", "annotator_2_label", "final_label"]


def read_table_flex(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
    except Exception:
        return pd.read_csv(path, encoding="utf-8-sig")


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def apply_excel_formatting(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = "Разметка"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    label_dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    label_dv.prompt = "Выберите метку: 0, 1 или 2"
    label_dv.error = "Допустимы только значения 0, 1, 2"
    ws.add_data_validation(label_dv)

    header_map = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    for col_name in LABEL_COLUMNS:
        if col_name in header_map and ws.max_row >= 2:
            col_idx = header_map[col_name]
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            label_dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    autosize_columns(ws)
    wb.save(xlsx_path)


def main() -> int:
    ANNOTATIONS_DIR.mkdir(parents=True, exist_ok=True)

    if not NEW_CSV.exists():
        raise FileNotFoundError(f"Не найден новый шаблон: {NEW_CSV}")

    new_df = read_table_flex(NEW_CSV)

    for col in KEY_COLUMNS:
        if col not in new_df.columns:
            raise ValueError(f"В новом шаблоне нет колонки: {col}")

    if OLD_FILLED_CSV.exists():
        old_df = read_table_flex(OLD_FILLED_CSV)
        for col in KEY_COLUMNS:
            if col not in old_df.columns:
                raise ValueError(f"В старой разметке нет колонки: {col}")

        keep_cols = KEY_COLUMNS + [c for c in ANNOTATION_COLUMNS if c in old_df.columns]
        old_small = old_df[keep_cols].copy().drop_duplicates(subset=KEY_COLUMNS, keep="last")

        merged = new_df.merge(old_small, on=KEY_COLUMNS, how="left", suffixes=("", "__old"))

        for col in ANNOTATION_COLUMNS:
            old_col = f"{col}__old"
            if old_col in merged.columns:
                if col not in merged.columns:
                    merged[col] = merged[old_col]
                else:
                    merged[col] = merged[col].where(
                        merged[col].notna() & (merged[col].astype(str).str.strip() != ""),
                        merged[old_col]
                    )
                merged = merged.drop(columns=[old_col])

        transferred = len(set(map(tuple, old_df[KEY_COLUMNS].astype(str).values.tolist())) &
                          set(map(tuple, new_df[KEY_COLUMNS].astype(str).values.tolist())))
    else:
        merged = new_df.copy()
        transferred = 0

    merged.to_csv(MERGED_CSV, index=False, encoding="utf-8-sig")
    merged.to_excel(MERGED_XLSX, index=False, sheet_name="Разметка")
    apply_excel_formatting(MERGED_XLSX)

    print(f"Готово: {MERGED_CSV}")
    print(f"Готово: {MERGED_XLSX}")
    print(f"Строк в шаблоне: {len(merged)}")
    print(f"Перенесено старых размеченных пар: {transferred}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
